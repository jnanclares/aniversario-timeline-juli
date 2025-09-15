import json
import os
import sys
import unicodedata
import re
from datetime import datetime, date, timezone
from typing import Any, Dict, List, Optional, Tuple


EXCEL_FILE = "Aniversario.xlsx"
PRE_DIR = "Pre"
POST_DIR = "Post"
OUTPUT_JSON = "timeline.json"
VALID_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}


def to_iso_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    # Try common string formats
    s = str(value).strip()
    if not s:
        return None
    # Replace common separators and try parsing
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return s  # last resort, return raw string


def read_excel_sheet(sheet_name: str) -> List[Dict[str, Any]]:
    # Attempt with pandas first for robustness, then fall back to openpyxl
    try:
        import pandas as pd  # type: ignore

        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        # Helper for robust header matching
        def normalize_header(s: str) -> str:
            s2 = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
            s2 = s2.lower().strip()
            for ch in ["#", "(", ")", ":", ";", ",", "."]:
                s2 = s2.replace(ch, " ")
            s2 = " ".join(s2.split())
            return s2

        original_headers = [str(c) for c in df.columns]
        norm_to_original: Dict[str, str] = {normalize_header(str(c)): str(c) for c in df.columns}

        def pick(name_variants: List[str]) -> Optional[str]:
            variants_norm = [normalize_header(v) for v in name_variants]
            # exact normalized match
            for v in variants_norm:
                if v in norm_to_original:
                    return norm_to_original[v]
            # contains or startswith matching
            for candidate_norm, original in norm_to_original.items():
                for v in variants_norm:
                    if v in candidate_norm or candidate_norm in v:
                        return original
            return None

        col_inicio = pick(["Inicio", "Start", "Fecha inicio", "Desde", "inicio (fecha)", "fecha inicio"])
        col_fin = pick(["Fin", "End", "Fecha fin", "Hasta", "fin (fecha)", "fecha fin"])
        col_semana = pick(["Semana", "Week", "Week #", "W", "semana (week)", "semana (week #)"])
        col_comentario = pick(["Comentario", "Comment", "Notas", "Descripción", "comentario (comentario)"])

        print(f"[{sheet_name}] Columnas detectadas: {original_headers}")
        print(f"[{sheet_name}] Mapeo usado -> Inicio: {col_inicio}, Fin: {col_fin}, Semana: {col_semana}, Comentario: {col_comentario}")

        # Diagnostics
        if col_semana is not None:
            sample_vals = [str(v) for v in list(df[col_semana].head(5).values) if v is not None]
            print(f"[{sheet_name}] Muestras Semana: {sample_vals}")

        def parse_week(value: Any) -> Optional[int]:
            if value is None:
                return None
            if isinstance(value, (int,)):
                return int(value)
            if isinstance(value, float):
                try:
                    return int(value)
                except Exception:
                    pass
            s = str(value).strip()
            if not s:
                return None
            m = re.search(r"(\d+)", s)
            if m:
                try:
                    return int(m.group(1))
                except Exception:
                    return None
            return None

        total_rows = len(df.index)
        non_empty_semana = 0
        conversion_failures = 0
        records: List[Dict[str, Any]] = []
        for _, row in df.iterrows():
            if col_semana is None:
                continue
            val = row.get(col_semana)
            if val is None or (hasattr(val, "__class__") and str(val) == "nan"):
                continue
            non_empty_semana += 1
            week_num = parse_week(val)
            if week_num is None:
                conversion_failures += 1
                continue
            records.append({
                "week": week_num,
                "start": to_iso_date(row.get(col_inicio)) if col_inicio else None,
                "end": to_iso_date(row.get(col_fin)) if col_fin else None,
                "comment": None if (col_comentario is None) else (None if (pd.isna(row.get(col_comentario))) else str(row.get(col_comentario))),
            })
        print(f"[{sheet_name}] Filas totales: {total_rows}, con Semana no vacía: {non_empty_semana}, conversiones fallidas: {conversion_failures}")
        print(f"[{sheet_name}] Filas parseadas: {len(records)}")
        return records
    except Exception:
        pass

    # Fallback: openpyxl
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(EXCEL_FILE, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        # Build header map
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in header_cells]
        def normalize_header(s: str) -> str:
            s2 = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
            s2 = s2.lower().strip()
            for ch in ["#", "(", ")", ":", ";", ",", "."]:
                s2 = s2.replace(ch, " ")
            s2 = " ".join(s2.split())
            return s2
        norm_headers = [normalize_header(h) for h in headers]
        def find_index(variants: List[str]) -> Optional[int]:
            variants_norm = [normalize_header(v) for v in variants]
            # exact match
            for v in variants_norm:
                if v in norm_headers:
                    return norm_headers.index(v)
            # contains
            for i, h in enumerate(norm_headers):
                for v in variants_norm:
                    if v in h or h in v:
                        return i
            return None

        idx_inicio = find_index(["Inicio", "Start", "Fecha inicio", "Desde"])
        idx_fin = find_index(["Fin", "End", "Fecha fin", "Hasta"])
        idx_semana = find_index(["Semana", "Week", "Week #", "W"])
        idx_comentario = find_index(["Comentario", "Comment", "Notas", "Descripción"])

        print(f"[{sheet_name}] Columnas detectadas: {headers}")
        print(f"[{sheet_name}] Índices usados -> Inicio: {idx_inicio}, Fin: {idx_fin}, Semana: {idx_semana}, Comentario: {idx_comentario}")

        def parse_week(value: Any) -> Optional[int]:
            if value is None:
                return None
            if isinstance(value, (int,)):
                return int(value)
            if isinstance(value, float):
                try:
                    return int(value)
                except Exception:
                    pass
            s = str(value).strip()
            if not s:
                return None
            m = re.search(r"(\d+)", s)
            if m:
                try:
                    return int(m.group(1))
                except Exception:
                    return None
            return None

        records = []
        total_rows = 0
        non_empty_semana = 0
        conversion_failures = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            if idx_semana is None:
                continue
            semana_val = row[idx_semana] if idx_semana < len(row) else None
            if semana_val is None:
                continue
            non_empty_semana += 1
            week_num = parse_week(semana_val)
            if week_num is None:
                conversion_failures += 1
                continue
            start_val = row[idx_inicio] if (idx_inicio is not None and idx_inicio < len(row)) else None
            end_val = row[idx_fin] if (idx_fin is not None and idx_fin < len(row)) else None
            comment_val = row[idx_comentario] if (idx_comentario is not None and idx_comentario < len(row)) else None
            records.append({
                "week": week_num,
                "start": to_iso_date(start_val),
                "end": to_iso_date(end_val),
                "comment": None if comment_val is None else str(comment_val),
            })
        print(f"[{sheet_name}] Filas totales: {total_rows}, con Semana no vacía: {non_empty_semana}, conversiones fallidas: {conversion_failures}")
        print(f"[{sheet_name}] Filas parseadas: {len(records)}")
        return records
    except Exception as e:
        print(f"Error leyendo {EXCEL_FILE} hoja {sheet_name}: {e}", file=sys.stderr)
        return []


def list_photos_for_week(base_dir: str, week_num: int) -> List[str]:
    week_dir = os.path.join(base_dir, f"W{week_num}")
    if not os.path.isdir(week_dir):
        return []
    files = []
    try:
        for name in os.listdir(week_dir):
            _, ext = os.path.splitext(name)
            if ext.lower() in VALID_EXTENSIONS:
                rel_path = os.path.join(base_dir, f"W{week_num}", name).replace("\\", "/")
                files.append(rel_path)
    except Exception:
        return []
    files.sort(key=lambda p: p.lower())
    return files


def merge_records_with_photos(records: List[Dict[str, Any]], base_dir: str) -> List[Dict[str, Any]]:
    merged = []
    for rec in records:
        week = rec.get("week")
        photos = list_photos_for_week(base_dir, week)
        merged.append({
            "week": week,
            "start": rec.get("start"),
            "end": rec.get("end"),
            "comment": rec.get("comment"),
            "photos": photos,
        })
    merged.sort(key=lambda r: r.get("week", 0))
    return merged


def main() -> int:
    if not os.path.exists(EXCEL_FILE):
        print(f"No se encontró {EXCEL_FILE} en el directorio actual.")
        return 1

    pre_records = read_excel_sheet("Pre")
    post_records = read_excel_sheet("Post")

    pre = merge_records_with_photos(pre_records, PRE_DIR)
    post = merge_records_with_photos(post_records, POST_DIR)

    payload = {
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "pre": pre,
        "post": post,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"Generado {OUTPUT_JSON} con {len(pre)} semanas Pre y {len(post)} semanas Post.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


